import io
from datetime import date
from typing import List, Dict, Any, Optional

from flask import (
    Blueprint, render_template, request, redirect,
    url_for, flash, send_file, jsonify
)
from flask_login import login_required, current_user
from msd.database.connection import get_conn

try:
    import openpyxl
except ImportError:
    openpyxl = None

employees_console_bp = Blueprint("employees_console_bp", __name__)

# ==========================
# إعدادات عامة
# ==========================
EXCEL_EXPECTED_ORDER = [
    "serial_number",   # A
    "name",            # B
    "national_id",     # C
    "hiring_date",     # D
    "job_grade",       # E
    "bonus",           # F
    "grade_date",      # G
    "vacation_balance",# H
    "department",      # I (اسم أو رقم)
    "work_days"        # J (اختياري)
]

CREATE_DEPARTMENT_IF_MISSING = True

WORK_DAYS_MAP = {
    "0": "السبت",
    "1": "الأحد",
    "2": "الإثنين",
    "3": "الثلاثاء",
    "4": "الأربعاء",
    "5": "الخميس",
    "6": "الجمعة"
}
PERIOD_MAP_AR = {"M": "صباحية", "E": "مسائية", "F": "كامل اليوم"}


# ==========================
# Utils
# ==========================
def decode_work_days(raw: Optional[str]) -> str:
    if not raw:
        return ""
    raw = raw.strip()
    if raw in ("الندب", "تفرغ"):
        return raw
    out = []
    parts = [p.strip() for p in raw.replace(";", ",").split(",") if p.strip()]
    for p in parts:
        if ":" in p:
            code, period = p.split(":", 1)
            day_ar = WORK_DAYS_MAP.get(code, code)
            per_ar = PERIOD_MAP_AR.get(period, "")
            if per_ar:
                out.append(f"{day_ar} ({per_ar})")
            else:
                out.append(day_ar)
        else:
            out.append(WORK_DAYS_MAP.get(p, p))
    return "، ".join(out)


def fetch_departments() -> List[Dict[str, Any]]:
    with get_conn() as conn:
        rows = conn.execute("SELECT id,name FROM departments ORDER BY name").fetchall()
    return [{"id": r["id"], "name": r["name"]} for r in rows]


def count_on_vacation() -> int:
    today = date.today().isoformat()
    sql = """
      SELECT COUNT(*) c
        FROM vacation_requests
       WHERE status='approved'
         AND date(start_date)<=date(?)
         AND date(end_date)>=date(?)
    """
    try:
        with get_conn() as conn:
            c = conn.execute(sql, (today, today)).fetchone()[0]
        return c or 0
    except Exception:
        return 0


def fetch_employees(limit=10000):
    sql = """
    SELECT e.id, e.serial_number, e.name, e.national_id,
           e.department_id, e.job_grade, e.hiring_date, e.grade_date,
           e.bonus,
           COALESCE(e.annual_balance, e.vacation_balance, 0) AS vacation_balance,
           COALESCE(e.status,'active') AS status,
           e.work_days,
           d.name AS dept_name
      FROM employees e
      LEFT JOIN departments d ON d.id = e.department_id
     ORDER BY e.id DESC
     LIMIT ?
    """
    with get_conn() as conn:
        rows = conn.execute(sql, (limit,)).fetchall()
    out = []
    for r in rows:
        out.append({
            "id": r["id"],
            "serial_number": r["serial_number"],
            "name": r["name"],
            "national_id": r["national_id"],
            "department_id": r["department_id"],
            "job_grade": r["job_grade"],
            "hiring_date": r["hiring_date"],
            "grade_date": r["grade_date"],
            "bonus": r["bonus"],
            "vacation_balance": r["vacation_balance"],
            "status": r["status"],
            "work_days": r["work_days"],
            "dept_name": r["dept_name"]
        })
    return out


def ensure_department(name_or_id) -> Optional[int]:
    """
    يحاول تفسير القيمة إما رقم (id موجود مسبقاً) أو اسم قسم.
    يعيد department_id أو None.
    ينشئ القسم إذا لم يوجد و CREATE_DEPARTMENT_IF_MISSING=True
    """
    if name_or_id in (None, ""):
        return None
    with get_conn() as conn:
        cur = conn.cursor()
        # رقم؟
        try:
            did = int(str(name_or_id).strip())
            row = cur.execute("SELECT id FROM departments WHERE id=?", (did,)).fetchone()
            if row:
                return did
        except Exception:
            pass
        # اسم
        dep_name = str(name_or_id).strip()
        row = cur.execute("SELECT id FROM departments WHERE name=?", (dep_name,)).fetchone()
        if row:
            return row[0]
        if CREATE_DEPARTMENT_IF_MISSING:
            cur.execute("INSERT INTO departments (name) VALUES (?)", (dep_name,))
            conn.commit()
            return cur.lastrowid
    return None


def validate_manager():
    if current_user.role not in ("manager", "admin"):
        return False
    return True


# ==========================
# Routes
# ==========================

@employees_console_bp.route("/manager/employees", methods=["GET"])
@login_required
def employees_console():
    if not validate_manager():
        return "غير مسموح", 403
    employees = fetch_employees()
    departments = fetch_departments()
    on_vac = count_on_vacation()
    return render_template(
        "manager.html",
        employees=employees,
        departments=departments,
        on_vacation_count=on_vac
    )


@employees_console_bp.post("/manager/employees/import")
@login_required
def employees_import():
    if not validate_manager():
        flash("غير مسموح", "danger")
        return redirect(url_for("employees_console_bp.employees_console"))
    if openpyxl is None:
        flash("مكتبة openpyxl غير مثبتة", "danger")
        return redirect(url_for("employees_console_bp.employees_console"))
    file = request.files.get("file")
    if not file:
        flash("لم يتم اختيار ملف", "warning")
        return redirect(url_for("employees_console_bp.employees_console"))
    try:
        data = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            flash("ملف فارغ", "warning")
            return redirect(url_for("employees_console_bp.employees_console"))
        # تحديد إن كان الصف الأول ترويسة
        header = [str(c).strip().lower() if c else "" for c in rows[0]]
        # نحاول اكتشاف إذا أول صف ترويسة
        start_index = 0
        if "name" in header or "national_id" in header or "serial_number" in header:
            start_index = 1

        inserted = 0
        skipped = 0
        with get_conn() as conn:
            cur = conn.cursor()
            for r in rows[start_index:]:
                if r is None:
                    continue
                # توحيد الطول
                row = list(r) + [None]* (len(EXCEL_EXPECTED_ORDER)-len(r))
                serial_number = str(row[0]).strip() if row[0] not in (None,"") else None
                name = str(row[1]).strip() if row[1] not in (None,"") else None
                national_id = str(row[2]).strip() if row[2] not in (None,"") else None
                hiring_date = str(row[3]).strip() if row[3] not in (None,"") else None
                job_grade = str(row[4]).strip() if row[4] not in (None,"") else None
                bonus = 0
                if row[5] not in (None,""):
                    try: bonus = int(float(row[5]))
                    except: bonus = 0
                grade_date = str(row[6]).strip() if row[6] not in (None,"") else None
                vacation_balance = 30
                if row[7] not in (None,""):
                    try: vacation_balance = float(row[7])
                    except: pass
                department_ref = row[8]
                work_days = row[9] if len(row) > 9 else None

                if not name:
                    skipped += 1
                    continue
                dept_id = ensure_department(department_ref)
                try:
                    cur.execute("""
                        INSERT INTO employees
                        (serial_number,name,national_id,department_id,job_grade,
                         hiring_date,grade_date,bonus,annual_balance,status,work_days)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?)
                    """, (
                        serial_number, name, national_id, dept_id, job_grade,
                        hiring_date, grade_date, bonus, vacation_balance, "active",
                        work_days
                    ))
                    inserted += 1
                except Exception:
                    skipped += 1
            conn.commit()
        flash(f"تم الاستيراد: {inserted} / تخطي: {skipped}", "success")
    except Exception as e:
        flash(f"فشل الاستيراد: {e}", "danger")
    return redirect(url_for("employees_console_bp.employees_console"))


@employees_console_bp.get("/manager/employees/export")
@login_required
def employees_export():
    if not validate_manager():
        return "غير مسموح", 403
    employees = fetch_employees()
    if openpyxl is None:
        # تصدير CSV بسيط
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(EXCEL_EXPECTED_ORDER)
        for e in employees:
            writer.writerow([
                e["serial_number"], e["name"], e["national_id"],
                e["hiring_date"], e["job_grade"], e["bonus"], e.get("grade_date"),
                e["vacation_balance"], e["dept_name"], e["work_days"]
            ])
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            mimetype="text/csv",
            as_attachment=True,
            download_name="employees_export.csv"
        )
    # XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.append(EXCEL_EXPECTED_ORDER)
    for e in employees:
        ws.append([
            e["serial_number"], e["name"], e["national_id"],
            e["hiring_date"], e["job_grade"], e["bonus"], e.get("grade_date"),
            e["vacation_balance"], e["dept_name"], e["work_days"]
        ])
    f = io.BytesIO()
    wb.save(f)
    f.seek(0)
    return send_file(
        f,
        as_attachment=True,
        download_name="employees_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@employees_console_bp.post("/manager/employees/add")
@login_required
def employees_add():
    if not validate_manager():
        return jsonify(success=False, error="غير مسموح"), 403
    data = request.get_json() or {}
    try:
        dept_id = ensure_department(data.get("department_id") or data.get("department"))
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO employees
                (serial_number,name,national_id,department_id,job_grade,
                 hiring_date,grade_date,bonus,annual_balance,status,work_days)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, (
                data.get("serial_number"),
                data.get("name"),
                data.get("national_id"),
                dept_id,
                data.get("job_grade"),
                data.get("hiring_date"),
                data.get("grade_date"),
                data.get("bonus") or 0,
                data.get("vacation_balance") or 30,
                "active",
                data.get("work_days")
            ))
            conn.commit()
        return jsonify(success=True, message="تمت الإضافة")
    except Exception as e:
        return jsonify(success=False, error=str(e)), 400


@employees_console_bp.put("/manager/employees/update/<int:emp_id>")
@login_required
def employees_update(emp_id):
    if not validate_manager():
        return jsonify(success=False, error="غير مسموح"), 403
    data = request.get_json() or {}
    try:
        dept_id = ensure_department(data.get("department_id") or data.get("department"))
        with get_conn() as conn:
            conn.execute("""
                UPDATE employees
                   SET serial_number=?, name=?, national_id=?, department_id=?, job_grade=?,
                       hiring_date=?, grade_date=?, bonus=?, annual_balance=?, work_days=?, status=?
                 WHERE id=?
            """, (
                data.get("serial_number"),
                data.get("name"),
                data.get("national_id"),
                dept_id,
                data.get("job_grade"),
                data.get("hiring_date"),
                data.get("grade_date"),
                data.get("bonus") or 0,
                data.get("vacation_balance") or 30,
                data.get("work_days"),
                data.get("status") or "active",
                emp_id
            ))
            conn.commit()
        return jsonify(success=True, message="تم التحديث")
    except Exception as e:
        return jsonify(success=False, error=str(e)), 400


@employees_console_bp.delete("/manager/employees/delete/<int:emp_id>")
@login_required
def employees_delete(emp_id):
    if not validate_manager():
        return jsonify(success=False, error="غير مسموح"), 403
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM employees WHERE id=?", (emp_id,))
            if cur.rowcount == 0:
                return jsonify(success=False, error="الموظف غير موجود"), 404
            conn.commit()
        return jsonify(success=True, message="تم الحذف")
    except Exception as e:
        return jsonify(success=False, error=str(e)), 400


@employees_console_bp.get("/manager/employees/<int:emp_id>/details")
@login_required
def employees_details(emp_id):
    if not validate_manager():
        return jsonify(success=False, error="غير مسموح"), 403
    try:
        with get_conn() as conn:
            emp = conn.execute("""
                SELECT e.id, e.serial_number, e.name, e.national_id,
                       e.department_id, d.name AS dept_name,
                       e.job_grade, e.hiring_date, e.grade_date,
                       e.bonus, COALESCE(e.annual_balance,e.vacation_balance,0) AS vacation_balance,
                       e.status, e.work_days
                  FROM employees e
                  LEFT JOIN departments d ON d.id=e.department_id
                 WHERE e.id=?
            """, (emp_id,)).fetchone()
            if not emp:
                return jsonify(success=False, error="غير موجود"), 404

            vacations = conn.execute("""
                SELECT id, type_code, start_date, end_date, requested_days, status
                  FROM vacation_requests
                 WHERE employee_id=?
                 ORDER BY id DESC LIMIT 10
            """, (emp_id,)).fetchall()

            absences = conn.execute("""
                SELECT id, type AS absence_type, start_date, end_date, duration
                  FROM absences
                 WHERE employee_id=?
                 ORDER BY id DESC LIMIT 10
            """, (emp_id,)).fetchall()

        emp_dict = {
            "id": emp["id"],
            "serial_number": emp["serial_number"],
            "name": emp["name"],
            "national_id": emp["national_id"],
            "department_id": emp["department_id"],
            "dept_name": emp["dept_name"],
            "job_grade": emp["job_grade"],
            "hiring_date": emp["hiring_date"],
            "grade_date": emp["grade_date"],
            "bonus": emp["bonus"],
            "vacation_balance": emp["vacation_balance"],
            "status": emp["status"],
            "work_days_raw": emp["work_days"],
            "work_days_human": decode_work_days(emp["work_days"])
        }

        vac_list = [{
            "id": v["id"],
            "type_code": v["type_code"],
            "start_date": v["start_date"],
            "end_date": v["end_date"],
            "requested_days": v["requested_days"],
            "status": v["status"]
        } for v in vacations]

        abs_list = [{
            "id": a["id"],
            "type": a["absence_type"],
            "start_date": a["start_date"],
            "end_date": a["end_date"],
            "duration": a["duration"]
        } for a in absences]

        return jsonify(success=True, employee=emp_dict, vacations=vac_list, absences=abs_list)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 400