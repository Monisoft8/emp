from datetime import date, datetime, timedelta
from typing import Optional, List, Dict, Tuple
from io import BytesIO
from msd.database.connection import get_conn

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font
except ImportError:
    openpyxl = None  # سنتحقق لاحقاً

# ---------------- Lazy Meta ----------------
# لا نحمل الأعمدة عند الاستيراد لتفادي الحاجة إلى Flask context
_META = {
    "cols": None,
    "has_range": False,
    "has_date": False,
    "loaded": False
}

TYPE_LABELS = {
    "absence": "غياب",
    "late": "تأخير",
    "early_leave": "انصراف مبكر"
}

def _ensure_meta():
    if _META["loaded"]:
        return
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(absences)")
        cols = {r[1] for r in cur.fetchall()}
    _META["cols"] = cols
    _META["has_range"] = "start_date" in cols and "end_date" in cols
    _META["has_date"] = "date" in cols
    _META["loaded"] = True

def _parse_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()

def _overlap_days(rec_start: date, rec_end: date, win_start: date, win_end: date) -> int:
    start = max(rec_start, win_start)
    end = min(rec_end, win_end)
    if end < start:
        return 0
    return (end - start).days + 1

def _fetch_candidate_records(win_start: date, win_end: date,
                             employee_id: Optional[int] = None) -> List[Dict]:
    _ensure_meta()
    where = []
    params = []
    if _META["has_range"]:
        where.append("NOT (end_date < ? OR start_date > ?)")
        params.extend([win_start.isoformat(), win_end.isoformat()])
    else:
        where.append("NOT (date < ? OR date > ?)")
        params.extend([win_start.isoformat(), win_end.isoformat()])

    if employee_id:
        where.append("a.employee_id=?")
        params.append(employee_id)

    where_sql = "WHERE " + " AND ".join(where) if where else ""
    range_select = "a.start_date, a.end_date" if _META["has_range"] else "a.date AS start_date, a.date AS end_date"

    sql = f"""
      SELECT a.id, a.employee_id, e.name AS employee_name,
             a.type, {range_select}, a.duration, a.notes
        FROM absences a
        LEFT JOIN employees e ON e.id=a.employee_id
        {where_sql}
      ORDER BY a.employee_id, a.type, a.start_date
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()
        return [dict(r) for r in rows]

def generate_report(report_type: str,
                    year: Optional[int] = None,
                    month: Optional[int] = None,
                    start_date: Optional[str] = None,
                    end_date: Optional[str] = None,
                    employee_id: Optional[int] = None) -> Dict:
    """
    أنواع التقرير:
      - month  (يستلزم year, month)
      - range  (start_date, end_date)
      - employee (employee_id) مع نطاق اختياري
    """
    if report_type == "month":
        if not (year and month):
            raise ValueError("يجب تحديد السنة والشهر")
        win_start = date(year, month, 1)
        if month == 12:
            win_end = date(year+1, 1, 1) - timedelta(days=1)
        else:
            win_end = date(year, month+1, 1) - timedelta(days=1)
    elif report_type == "range":
        if not (start_date and end_date):
            raise ValueError("يجب تحديد تاريخي البداية والنهاية")
        win_start = _parse_date(start_date)
        win_end = _parse_date(end_date)
        if win_end < win_start:
            raise ValueError("نهاية الفترة قبل بدايتها")
    elif report_type == "employee":
        if not employee_id:
            raise ValueError("يجب تحديد الموظف للتقرير الخاص")
        if start_date and end_date:
            win_start = _parse_date(start_date)
            win_end = _parse_date(end_date)
            if win_end < win_start:
                raise ValueError("نهاية الفترة قبل بدايتها")
        else:
            win_start = date(2000,1,1)
            win_end = date(2100,1,1)
    else:
        raise ValueError("نوع تقرير غير مدعوم")

    candidates = _fetch_candidate_records(win_start, win_end,
                                          employee_id if report_type=="employee" else None)

    aggregated = {}
    per_employee_total = {}
    grand_total = 0

    for rec in candidates:
        try:
            r_start = _parse_date(rec['start_date'])
            r_end = _parse_date(rec['end_date'])
        except Exception:
            continue
        overlap = _overlap_days(r_start, r_end, win_start, win_end)
        if overlap <= 0:
            continue
        if rec['type'] in ("late", "early_leave"):
            overlap = 1

        key = (rec['employee_id'], rec['type'])
        if key not in aggregated:
            aggregated[key] = {
                "employee_id": rec['employee_id'],
                "employee_name": rec.get('employee_name') or f"#{rec['employee_id']}",
                "type": rec['type'],
                "type_label": TYPE_LABELS.get(rec['type'], rec['type']),
                "days": 0
            }
        aggregated[key]["days"] += overlap
        per_employee_total.setdefault(rec['employee_id'], 0)
        per_employee_total[rec['employee_id']] += overlap
        grand_total += overlap

    rows = list(aggregated.values())
    rows.sort(key=lambda r: (r["employee_name"], r["type"]))

    return {
        "params": {
            "report_type": report_type,
            "year": year,
            "month": month,
            "start_date": win_start.isoformat(),
            "end_date": win_end.isoformat(),
            "employee_id": employee_id
        },
        "items": rows,
        "totals": {
            "grand_total_days": grand_total,
            "per_employee": per_employee_total
        }
    }

def export_report_to_excel(report_data: Dict) -> Tuple[bytes, str, str]:
    if openpyxl is None:
        raise RuntimeError("مكتبة openpyxl غير مثبتة (pip install openpyxl)")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    params = report_data.get("params", {})
    items = report_data.get("items", [])
    totals = report_data.get("totals", {})

    title = f"تقرير الغياب - نوع: {params.get('report_type')}"
    ws.merge_cells("A1:C1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append(["النطاق", f"{params.get('start_date')} → {params.get('end_date')}"])
    ws.append(["عدد الصفوف المجمعة", len(items)])
    ws.append([])

    ws.append(["الموظف","النوع","الأيام"])
    ws.row_dimensions[ws.max_row].font = Font(bold=True)

    for r in items:
        ws.append([
            r["employee_name"],
            r["type_label"],
            r["days"]
        ])

    ws.append([])
    ws.append(["الإجمالي الكلي للأيام", totals.get("grand_total_days", 0)])

    if totals.get("per_employee"):
        ws.append([])
        ws.append(["إجمالي لكل موظف (ID, Days)"])
        ws.row_dimensions[ws.max_row].font = Font(bold=True)
        for emp_id, d in totals["per_employee"].items():
            ws.append([emp_id, d])

    for col in ("A","B","C","D","E"):
        if ws.column_dimensions.get(col):
            ws.column_dimensions[col].width = 28

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"absences_report_{params.get('report_type')}.xlsx"
    return bio.getvalue(), filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"